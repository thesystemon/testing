
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment



# Module 1 (Basic Features)

def read_keywords_from_file(file_path):
    with open(file_path, 'r') as file:
        keywords = file.read().strip().split(',')
    return [keyword.strip() for keyword in keywords]

def read_and_process_csv(csv_files, keywords):
    combined_data = None

    # Combine keywords into a single regex pattern
    keyword_pattern = '|'.join(keywords)

    for csv_file in csv_files:
        data = pd.read_csv(csv_file)
        # Filter data to include only rows where any of the keywords are present in the label
        filtered_data = data[data['label'].str.contains(keyword_pattern, case=False, na=False)]
        
        grouped_data = filtered_data.groupby('label').agg(
            count=('elapsed', 'count'),
            average=('elapsed', 'mean'),
            median=('elapsed', 'median'),
            average_90_per=('elapsed', lambda x: x.quantile(0.90))
        ).reset_index()

        # Sorting priority based on multiple keywords
        def label_priority(label):
            for i, keyword in enumerate(keywords):
                if keyword.lower() in label.lower():
                    return i
            return len(keywords)
        
        grouped_data['label_priority'] = grouped_data['label'].apply(label_priority)
        grouped_data = grouped_data.sort_values(by=['label_priority', 'label']).drop(columns=['label_priority'])

        prefix = os.path.splitext(os.path.basename(csv_file))[0]
        grouped_data.columns = ['Label'] + [f'{prefix}_{col}' for col in grouped_data.columns[1:]]

        if combined_data is None:
            combined_data = grouped_data
        else:
            combined_data = pd.merge(combined_data, grouped_data, on='Label', how='outer')

    return combined_data

def create_excel_workbook(combined_data, csv_files):
    wb = Workbook()
    ws = wb.active
    ws.title = "Transaction Comparison"

    # Define styling elements
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    title_font = Font(bold=True, size=12)

    # Write "Transaction name" in cell A1
    ws.cell(row=2, column=1, value="Transaction name").font = title_font

    # Write headers and file names
    col_idx = 2
    for i, csv_file in enumerate(csv_files):
        ws.merge_cells(start_row=1, start_column=col_idx, end_row=1, end_column=col_idx + 3)
        merge_cell = ws.cell(row=1, column=col_idx)
        merge_cell.value = csv_file
        merge_cell.font = title_font
        merge_cell.alignment = Alignment(horizontal="center", vertical="center")

        headers = ['Count', 'Average', 'Median', '90 Percentile']
        for c_idx, col_name in enumerate(headers, start=col_idx):
            cell = ws.cell(row=2, column=c_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        col_idx += 4

    # Write data
    for r_idx, row in enumerate(combined_data.values, start=3):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    # Create charts
    create_charts(ws, combined_data, csv_files)

    # Create Deviation sheet
    ws_deviation = wb.create_sheet(title="Deviation")

    # Merge cells B1 to I1 and name it as "Deviation-25U_Revn_1404 & 25U_Revn_1405"
    ws_deviation.merge_cells('B1:I1')
    deviation_title_cell = ws_deviation.cell(row=1, column=2)
    deviation_title_cell.value = "Deviation-25U_Revn_1404 & 25U_Revn_1405"
    deviation_title_cell.font = title_font
    deviation_title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Merge cells B2 to E2 and name it "Deviation"
    ws_deviation.merge_cells('B2:E2')
    deviation_cell = ws_deviation.cell(row=2, column=2)
    deviation_cell.value = "Deviation"
    deviation_cell.font = title_font
    deviation_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Merge cells F2 to I2 and name it "Deviation %"
    ws_deviation.merge_cells('F2:I2')
    ws_deviation['F2'] = "Deviation %"
    ws_deviation['F2'].font = title_font
    ws_deviation['F2'].alignment = Alignment(horizontal="center", vertical="center")

    # Copy the second row headers from the first sheet
    for c_idx, cell in enumerate(ws[2], start=1):
        new_cell = ws_deviation.cell(row=3, column=c_idx, value=cell.value)
        new_cell.font = header_font
        new_cell.fill = header_fill
        new_cell.alignment = header_alignment

    # Calculate and write differences and deviations if count columns exist for both CSV files
    label_col = 'Label'
    count_cols = [col for col in combined_data.columns if 'count' in col]
    average_cols = [col for col in combined_data.columns if 'average' in col]
    median_cols = [col for col in combined_data.columns if 'median' in col]
    percentile_cols = [col for col in combined_data.columns if 'average_90_per' in col]

    if len(count_cols) == len(csv_files):
        for r_idx, label in enumerate(combined_data[label_col], start=4):
            count_diff = combined_data.loc[combined_data[label_col] == label, count_cols[-1]].values[0] \
                         - combined_data.loc[combined_data[label_col] == label, count_cols[-2]].values[0]
            average_diff = combined_data.loc[combined_data[label_col] == label, f'{csv_files[-1].split(".")[0]}_average'].values[0] \
                           - combined_data.loc[combined_data[label_col] == label, f'{csv_files[-2].split(".")[0]}_average'].values[0]
            median_diff = combined_data.loc[combined_data[label_col] == label, f'{csv_files[-1].split(".")[0]}_median'].values[0] \
                          - combined_data.loc[combined_data[label_col] == label, f'{csv_files[-2].split(".")[0]}_median'].values[0]
            percentile_diff = combined_data.loc[combined_data[label_col] == label, f'{csv_files[-1].split(".")[0]}_average_90_per'].values[0] \
                              - combined_data.loc[combined_data[label_col] == label, f'{csv_files[-2].split(".")[0]}_average_90_per'].values[0]
            ws_deviation.cell(row=r_idx, column=1, value=label)
            ws_deviation.cell(row=r_idx, column=2, value=count_diff)
            ws_deviation.cell(row=r_idx, column=3, value=average_diff)
            ws_deviation.cell(row=r_idx, column=4, value=median_diff)
            ws_deviation.cell(row=r_idx, column=5, value=percentile_diff)

            # Calculate and write percentage deviation for count
            count_previous = combined_data.loc[combined_data[label_col] == label, count_cols[-2]].values[0]
            if count_previous != 0:  # Avoid division by zero
                percentage_deviation_count = (count_diff / count_previous) * 100
                percentage_cell_count = ws_deviation.cell(row=r_idx, column=6, value=percentage_deviation_count / 100)
                percentage_cell_count.number_format = '0.00%'

            # Calculate and write percentage deviation for average
            average_previous = combined_data.loc[combined_data[label_col] == label, f'{csv_files[-2].split(".")[0]}_average'].values[0]
            if average_previous != 0:  # Avoid division by zero
                percentage_deviation_average = (average_diff / average_previous) * 100
                percentage_cell_average = ws_deviation.cell(row=r_idx, column=7, value=percentage_deviation_average / 100)
                percentage_cell_average.number_format = '0.00%'
            else:
                percentage_cell_average = ws_deviation.cell(row=r_idx, column=7, value=0)
                percentage_cell_average.number_format = '0.00%'

            # Calculate and write percentage deviation for median
            median_previous = combined_data.loc[combined_data[label_col] == label, f'{csv_files[-2].split(".")[0]}_median'].values[0]
            if median_previous != 0:  # Avoid division by zero
                percentage_deviation_median = (median_diff / median_previous) * 100
                percentage_cell_median = ws_deviation.cell(row=r_idx, column=8, value=percentage_deviation_median / 100)
                percentage_cell_median.number_format = '0.00%'
            else:
                percentage_cell_median = ws_deviation.cell(row=r_idx, column=8, value=0)
                percentage_cell_median.number_format = '0.00%'

            # Calculate and write percentage deviation for 90th percentile
            percentile_previous = combined_data.loc[combined_data[label_col] == label, f'{csv_files[-2].split(".")[0]}_average_90_per'].values[0]
            if percentile_previous != 0:  # Avoid division by zero
                percentage_deviation_percentile = (percentile_diff / percentile_previous) * 100
                percentage_cell_percentile = ws_deviation.cell(row=r_idx, column=9, value=percentage_deviation_percentile / 100)
                percentage_cell_percentile.number_format = '0.00%'
            else:
                percentage_cell_percentile = ws_deviation.cell(row=r_idx, column=9, value=0)
                percentage_cell_percentile.number_format = '0.00%'

    return wb

def create_charts(ws, combined_data, csv_files):
    chart_titles = ['Elapsed Time Metrics - Count by Label', 
                    'Elapsed Time Metrics - Average by Label', 
                    'Elapsed Time Metrics - Median by Label', 
                    'Elapsed Time Metrics - 90th Percentile by Label']
    chart_data_cols = [(2, 4), (3, 4), (4, 4), (5, 4)]

    for idx, (title, (data_col, num_cols)) in enumerate(zip(chart_titles, chart_data_cols)):
        chart = BarChart()
        chart.type = "col"
        chart.style = 10 + idx
        chart.title = title
        chart.y_axis.title = title.split(' - ')[-1]
        chart.x_axis.title = 'Label'
        chart.legend.position = "b"
        chart.width = 20
        chart.height = 15

        for i, csv_file in enumerate(csv_files):
            data_range = Reference(ws, min_col=data_col + 4 * i, min_row=2, max_row=len(combined_data) + 2)
            chart.add_data(data_range, titles_from_data=True)

        categories = Reference(ws, min_col=1, min_row=3, max_row=len(combined_data) + 2)
        chart.set_categories(categories)

        ws.add_chart(chart, f"A{len(combined_data) * (idx + 1) + 3}")

# Main Script
def main():
    # Option 1: Read keywords from a text file
    keywords = read_keywords_from_file('keywords.txt')
    
    # Specify the paths to the directories containing CSV files
    folder_path_basic = 'C:\\Users\\vyank\\OneDrive\\Desktop\\Performance-auto\\Files_Basic'


    # Get all CSV files from the directories
    csv_files_basic = [os.path.join(folder_path_basic, file) for file in os.listdir(folder_path_basic) if file.endswith('.csv')]

    #output are save 

    combined_data = read_and_process_csv(csv_files_basic, keywords)
    wb = create_excel_workbook(combined_data, [os.path.basename(file) for file in csv_files_basic])
    wb.save("basic=module-result.xlsx")
    print(f'Processed data saved to basic=module-result.xlsx')         


if __name__ == "__main__":
    main()
