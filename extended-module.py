
# import os
# import pandas as pd
# from openpyxl import Workbook
# from openpyxl.chart import BarChart, Reference
# from openpyxl.styles import Font, PatternFill, Alignment


# def read_keywords_from_file(file_path):
#     with open(file_path, 'r') as file:
#         keywords = file.read().strip().split(',')
#     return [keyword.strip() for keyword in keywords]


# def read_and_process_csv(csv_files, keywords):
#     combined_data = None

#     keyword_pattern = '|'.join(keywords)

#     for csv_file in csv_files:
#         data = pd.read_csv(csv_file)
#         filtered_data = data[data['label'].str.contains(keyword_pattern, case=False, na=False)]
        
#         grouped_data = filtered_data.groupby('label').agg(
#             count=('elapsed', 'count'),
#             average=('elapsed', 'mean'),
#             median=('elapsed', 'median'),
#             average_90_per=('elapsed', lambda x: x.quantile(0.90))
#         ).reset_index()

#         def label_priority(label):
#             for i, keyword in enumerate(keywords):
#                 if keyword.lower() in label.lower():
#                     return i
#             return len(keywords)
        
#         grouped_data['label_priority'] = grouped_data['label'].apply(label_priority)
#         grouped_data = grouped_data.sort_values(by=['label_priority', 'label']).drop(columns=['label_priority'])

#         prefix = os.path.splitext(os.path.basename(csv_file))[0]
#         grouped_data.columns = ['Label'] + [f'{prefix}_{col}' for col in grouped_data.columns[1:]]

#         if combined_data is None:
#             combined_data = grouped_data
#         else:
#             combined_data = pd.merge(combined_data, grouped_data, on='Label', how='outer')

#     return combined_data


# def create_charts(ws, combined_data, csv_files):
#     chart_titles = ['Elapsed Time Metrics - Count by Label', 
#                     'Elapsed Time Metrics - Average by Label', 
#                     'Elapsed Time Metrics - Median by Label', 
#                     'Elapsed Time Metrics - 90th Percentile by Label']
#     chart_data_cols = [(2, 4), (3, 4), (4, 4), (5, 4)]

#     for idx, (title, (data_col, num_cols)) in enumerate(zip(chart_titles, chart_data_cols)):
#         chart = BarChart()
#         chart.type = "col"
#         chart.style = 10 + idx
#         chart.title = title
#         chart.y_axis.title = title.split(' - ')[-1]
#         chart.x_axis.title = 'Label'
#         chart.legend.position = "b"
#         chart.width = 20
#         chart.height = 15

#         for i, csv_file in enumerate(csv_files):
#             data_range = Reference(ws, min_col=data_col + 4 * i, min_row=3, max_row=len(combined_data) + 3)
#             chart.add_data(data_range, titles_from_data=True)

#         categories = Reference(ws, min_col=1, min_row=3, max_row=len(combined_data) + 3)
#         chart.set_categories(categories)

#         ws.add_chart(chart, f"A{len(combined_data) * (idx + 1) + 3}")

# def create_excel_workbook_extended(combined_data, csv_files):
#     wb = Workbook()
#     ws = wb.active
#     ws.title = "Transaction Comparison"

#     header_font = Font(bold=True, color="FFFFFF")
#     header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
#     header_alignment = Alignment(horizontal="center", vertical="center")
#     title_font = Font(bold=True, size=12)

#     # Add a row and merge cells from B1 to Q1, and from R1 to AG1
#     ws.insert_rows(1)
#     ws.merge_cells('B1:Q1')
#     cycle_1_cell = ws.cell(row=1, column=2)
#     cycle_1_cell.value = "Cycle 1"
#     cycle_1_cell.font = title_font
#     cycle_1_cell.alignment = Alignment(horizontal="center", vertical="center")
    
#     ws.merge_cells('R1:AG1')
#     cycle_2_cell = ws.cell(row=1, column=18)
#     cycle_2_cell.value = "Cycle 2"
#     cycle_2_cell.font = title_font
#     cycle_2_cell.alignment = Alignment(horizontal="center", vertical="center")

#     ws.cell(row=2, column=1, value="Transaction name").font = title_font

#     col_idx = 2
#     for i, csv_file in enumerate(csv_files):
#         ws.merge_cells(start_row=2, start_column=col_idx, end_row=2, end_column=col_idx + 3)
#         merge_cell = ws.cell(row=2, column=col_idx)
#         merge_cell.value = csv_file
#         merge_cell.font = title_font
#         merge_cell.alignment = Alignment(horizontal="center", vertical="center")

#         headers = ['Count', 'Average', 'Median', '90 Percentile']
#         for c_idx, col_name in enumerate(headers, start=col_idx):
#             cell = ws.cell(row=3, column=c_idx, value=col_name)
#             cell.font = header_font
#             cell.fill = header_fill
#             cell.alignment = header_alignment
#         col_idx += 4

#     for r_idx, row in enumerate(combined_data.values, start=4):  # Start from row 4
#         for c_idx, value in enumerate(row, start=1):
#             ws.cell(row=r_idx, column=c_idx, value=value)

#     create_charts(ws, combined_data, csv_files)  # Adjusted to reflect the new row

#     ws_deviation = wb.create_sheet(title="Deviation")

#     # Merge cells and format the headers
#     ws_deviation.merge_cells('B1:E1')
#     cycle_1_cell = ws_deviation.cell(row=1, column=2)
#     cycle_1_cell.value = "Cycle 1"
#     cycle_1_cell.font = title_font
#     cycle_1_cell.alignment = Alignment(horizontal="center", vertical="center")

#     ws_deviation.merge_cells('F1:I1')
#     cycle_2_cell = ws_deviation.cell(row=1, column=6)
#     cycle_2_cell.value = "Cycle 2"
#     cycle_2_cell.font = title_font
#     cycle_2_cell.alignment = Alignment(horizontal="center", vertical="center")

#     ws_deviation.merge_cells('J1:M1')
#     deviation_cell = ws_deviation.cell(row=1, column=10)
#     deviation_cell.value = "Deviation"
#     deviation_cell.font = title_font
#     deviation_cell.alignment = Alignment(horizontal="center", vertical="center")

#     ws_deviation.merge_cells('N1:Q1')
#     deviation_percent_cell = ws_deviation.cell(row=1, column=14)
#     deviation_percent_cell.value = "Deviation%"
#     deviation_percent_cell.font = title_font
#     deviation_percent_cell.alignment = Alignment(horizontal="center", vertical="center")

#     # Add headers for Cycle 1, Cycle 2, Deviation, and Deviation%
#     ws_deviation.cell(row=2, column=1, value="Transaction name").font = title_font
#     headers = ['Count', 'Average', 'Median', '90 Percentile']
#     for i, header in enumerate(headers, start=2):
#         cell_cycle_1 = ws_deviation.cell(row=2, column=i, value=header)
#         cell_cycle_1.font = header_font
#         cell_cycle_1.fill = header_fill
#         cell_cycle_1.alignment = header_alignment

#         cell_cycle_2 = ws_deviation.cell(row=2, column=i + 4, value=header)
#         cell_cycle_2.font = header_font
#         cell_cycle_2.fill = header_fill
#         cell_cycle_2.alignment = header_alignment

#         cell_deviation = ws_deviation.cell(row=2, column=i + 8, value=header)
#         cell_deviation.font = header_font
#         cell_deviation.fill = header_fill
#         cell_deviation.alignment = header_alignment

#         cell_deviation_percent = ws_deviation.cell(row=2, column=i + 12, value=header)
#         cell_deviation_percent.font = header_font
#         cell_deviation_percent.fill = header_fill
#         cell_deviation_percent.alignment = header_alignment

#     for r_idx, transaction_name in enumerate(ws.iter_rows(min_row=4, max_row=len(combined_data) + 3, min_col=1, max_col=1, values_only=True), start=4):
#         ws_deviation.cell(row=r_idx, column=1, value=transaction_name[0])
    
#     for r_idx in range(4, len(combined_data) + 4):  # Adjusted to start from row 4
#         cycle_1_values = []
#         cycle_2_values = []

#         for i in range(4):
#             cycle_1_values.append([
#                 ws.cell(row=r_idx, column=2 + i * 4).value,
#                 ws.cell(row=r_idx, column=3 + i * 4).value,
#                 ws.cell(row=r_idx, column=4 + i * 4).value,
#                 ws.cell(row=r_idx, column=5 + i * 4).value
#             ])
#             cycle_2_values.append([
#                 ws.cell(row=r_idx, column=18 + i * 4).value,
#                 ws.cell(row=r_idx, column=19 + i * 4).value,
#                 ws.cell(row=r_idx, column=20 + i * 4).value,
#                 ws.cell(row=r_idx, column=21 + i * 4).value
#             ])

#         for idx in range(4):
#             cycle_1_avg = [val for val in list(zip(*cycle_1_values))[idx] if isinstance(val, (int, float))]
#             cycle_2_avg = [val for val in list(zip(*cycle_2_values))[idx] if isinstance(val, (int, float))]
            
#             if cycle_1_avg:
#                 ws_deviation.cell(row=r_idx, column=2 + idx, value=sum(cycle_1_avg) / len(cycle_1_avg))
#             if cycle_2_avg:
#                 ws_deviation.cell(row=r_idx, column=6 + idx, value=sum(cycle_2_avg) / len(cycle_2_avg))

#             # Calculate deviation and deviation%
#             if cycle_1_avg and cycle_2_avg:
#                 deviation = (sum(cycle_2_avg) / len(cycle_2_avg)) - (sum(cycle_1_avg) / len(cycle_1_avg))
#                 deviation_percent = (deviation / (sum(cycle_1_avg) / len(cycle_1_avg))) * 100 if sum(cycle_1_avg) / len(cycle_1_avg) != 0 else 0
#                 deviation_percent_rounded = round(deviation_percent, 2)  # Round to 2 decimal places
#                 deviation_percent_str = f"{deviation_percent_rounded}%"  # Add % sign
#                 ws_deviation.cell(row=r_idx, column=10 + idx, value=deviation)
#                 ws_deviation.cell(row=r_idx, column=14 + idx, value=deviation_percent_str)

#     return wb





# def main():
#     keywords = read_keywords_from_file('keywords.txt')
    
    
#     folder_path_extended = 'C:\\Users\\vyank\\OneDrive\\Desktop\\Performance-auto\\Files_Extended'

    
#     csv_files_extended = [os.path.join(folder_path_extended, file) for file in os.listdir(folder_path_extended) if file.endswith('.csv')]
#     combined_data = read_and_process_csv(csv_files_extended, keywords)
#     wb = create_excel_workbook_extended(combined_data, [os.path.basename(file) for file in csv_files_extended])
#     output_file_extended = "extended-module-result.xlsx"
#     wb.save(output_file_extended)
#     print(f'Processed data saved to extended-module-result.xlsx')

        
# if __name__ == "__main__":
#     main()








import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment


def read_keywords_from_file(file_path):
    with open(file_path, 'r') as file:
        keywords = file.read().strip().split(',')
    return [keyword.strip() for keyword in keywords]


def read_and_process_csv(csv_files, keywords):
    combined_data = None

    keyword_pattern = '|'.join(keywords)

    for csv_file in csv_files:
        data = pd.read_csv(csv_file)
        filtered_data = data[data['label'].str.contains(keyword_pattern, case=False, na=False)]
        
        grouped_data = filtered_data.groupby('label').agg(
            count=('elapsed', 'count'),
            average=('elapsed', 'mean'),
            median=('elapsed', 'median'),
            average_90_per=('elapsed', lambda x: x.quantile(0.90))
        ).reset_index()

        def label_priority(label):
            for i, keyword in enumerate(keywords):
                if keyword.lower() in label.lower():
                    return i
            return len(keywords)
        
        grouped_data['label_priority'] = grouped_data['label'].apply(label_priority)
        grouped_data = grouped_data.sort_values(by=['label_priority', 'label']).drop(columns=['label_priority'])

        prefix = os.path.splitext(os.path.basename(csv_file))[0]
        grouped_data.columns = ['Label'] + [f'{prefix}_{col}' for col in grouped_data.columns[1:]]

        if combined_data is None:
            combined_data = grouped_data
        else:
            combined_data = pd.merge(combined_data, grouped_data, on='Label', how='outer')

    return combined_data


def create_charts(ws, combined_data, csv_files):
    chart_titles = ['Elapsed Time Metrics - Count by Label', 
                    'Elapsed Time Metrics - Average by Label', 
                    'Elapsed Time Metrics - Median by Label', 
                    'Elapsed Time Metrics - 90th Percentile by Label']
    chart_data_cols = [(2, 4), (3, 4), (4, 4), (5, 4)]

    for idx, (title, (data_col, num_cols)) in enumerate(zip(chart_titles, chart_data_cols)):
        chart = BarChart()
        chart.type = "col"
        chart.style = 10 + idx
        chart.title = title
        chart.y_axis.title = title.split(' - ')[-1]
        chart.x_axis.title = 'Label'
        chart.legend.position = "b"
        chart.width = 20
        chart.height = 15

        for i, csv_file in enumerate(csv_files):
            data_range = Reference(ws, min_col=data_col + 4 * i, min_row=3, max_row=len(combined_data) + 3)
            chart.add_data(data_range, titles_from_data=True)

        categories = Reference(ws, min_col=1, min_row=3, max_row=len(combined_data) + 3)
        chart.set_categories(categories)

        ws.add_chart(chart, f"A{len(combined_data) * (idx + 1) + 3}")

def create_excel_workbook_extended(combined_data, csv_files):
    wb = Workbook()
    ws = wb.active
    ws.title = "Transaction Comparison"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    title_font = Font(bold=True, size=12)

    # Add a row and merge cells from B1 to Q1, and from R1 to AG1
    ws.insert_rows(1)
    ws.merge_cells('B1:Q1')
    cycle_1_cell = ws.cell(row=1, column=2)
    cycle_1_cell.value = "Cycle 1"
    cycle_1_cell.font = title_font
    cycle_1_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    ws.merge_cells('R1:AG1')
    cycle_2_cell = ws.cell(row=1, column=18)
    cycle_2_cell.value = "Cycle 2"
    cycle_2_cell.font = title_font
    cycle_2_cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.cell(row=2, column=1, value="Transaction name").font = title_font

    col_idx = 2
    for i, csv_file in enumerate(csv_files):
        ws.merge_cells(start_row=2, start_column=col_idx, end_row=2, end_column=col_idx + 3)
        merge_cell = ws.cell(row=2, column=col_idx)
        merge_cell.value = csv_file
        merge_cell.font = title_font
        merge_cell.alignment = Alignment(horizontal="center", vertical="center")

        headers = ['Count', 'Average', 'Median', '90 Percentile']
        for c_idx, col_name in enumerate(headers, start=col_idx):
            cell = ws.cell(row=3, column=c_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        col_idx += 4

    for r_idx, row in enumerate(combined_data.values, start=4):  # Start from row 4
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    create_charts(ws, combined_data, csv_files)  # Adjusted to reflect the new row

    ws_deviation = wb.create_sheet(title="Deviation")

    # Merge cells and format the headers
    ws_deviation.merge_cells('B1:E1')
    cycle_1_cell = ws_deviation.cell(row=1, column=2)
    cycle_1_cell.value = "Cycle 1"
    cycle_1_cell.font = title_font
    cycle_1_cell.alignment = Alignment(horizontal="center", vertical="center")

    ws_deviation.merge_cells('F1:I1')
    cycle_2_cell = ws_deviation.cell(row=1, column=6)
    cycle_2_cell.value = "Cycle 2"
    cycle_2_cell.font = title_font
    cycle_2_cell.alignment = Alignment(horizontal="center", vertical="center")

    ws_deviation.merge_cells('J1:M1')
    deviation_cell = ws_deviation.cell(row=1, column=10)
    deviation_cell.value = "Deviation"
    deviation_cell.font = title_font
    deviation_cell.alignment = Alignment(horizontal="center", vertical="center")

    ws_deviation.merge_cells('N1:Q1')
    deviation_percent_cell = ws_deviation.cell(row=1, column=14)
    deviation_percent_cell.value = "Deviation%"
    deviation_percent_cell.font = title_font
    deviation_percent_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Add headers for Cycle 1, Cycle 2, Deviation, and Deviation%
    ws_deviation.cell(row=2, column=1, value="Transaction name").font = title_font
    headers = ['Count', 'Average', 'Median', '90 Percentile']
    for i, header in enumerate(headers, start=2):
        cell_cycle_1 = ws_deviation.cell(row=2, column=i, value=header)
        cell_cycle_1.font = header_font
        cell_cycle_1.fill = header_fill
        cell_cycle_1.alignment = header_alignment

        cell_cycle_2 = ws_deviation.cell(row=2, column=i + 4, value=header)
        cell_cycle_2.font = header_font
        cell_cycle_2.fill = header_fill
        cell_cycle_2.alignment = header_alignment

        cell_deviation = ws_deviation.cell(row=2, column=i + 8, value=header)
        cell_deviation.font = header_font
        cell_deviation.fill = header_fill
        cell_deviation.alignment = header_alignment

        cell_deviation_percent = ws_deviation.cell(row=2, column=i + 12, value=header)
        cell_deviation_percent.font = header_font
        cell_deviation_percent.fill = header_fill
        cell_deviation_percent.alignment = header_alignment

    for r_idx, transaction_name in enumerate(ws.iter_rows(min_row=4, max_row=len(combined_data) + 3, min_col=1, max_col=1, values_only=True), start=4):
        ws_deviation.cell(row=r_idx, column=1, value=transaction_name[0])
    
    for r_idx in range(4, len(combined_data) + 4):  # Adjusted to start from row 4
        cycle_1_values = []
        cycle_2_values = []

        for i in range(4):
            cycle_1_values.append([
                ws.cell(row=r_idx, column=2 + i * 4).value,
                ws.cell(row=r_idx, column=3 + i * 4).value,
                ws.cell(row=r_idx, column=4 + i * 4).value,
                ws.cell(row=r_idx, column=5 + i * 4).value
            ])
            cycle_2_values.append([
                ws.cell(row=r_idx, column=18 + i * 4).value,
                ws.cell(row=r_idx, column=19 + i * 4).value,
                ws.cell(row=r_idx, column=20 + i * 4).value,
                ws.cell(row=r_idx, column=21 + i * 4).value
            ])

        for idx in range(4):
            cycle_1_avg = [val for val in list(zip(*cycle_1_values))[idx] if isinstance(val, (int, float))]
            cycle_2_avg = [val for val in list(zip(*cycle_2_values))[idx] if isinstance(val, (int, float))]
            
            if cycle_1_avg:
                ws_deviation.cell(row=r_idx, column=2 + idx, value=sum(cycle_1_avg) / len(cycle_1_avg))
            if cycle_2_avg:
                ws_deviation.cell(row=r_idx, column=6 + idx, value=sum(cycle_2_avg) / len(cycle_2_avg))

            # Calculate deviation and deviation%
            if cycle_1_avg and cycle_2_avg:
                deviation = (sum(cycle_2_avg) / len(cycle_2_avg)) - (sum(cycle_1_avg) / len(cycle_1_avg))
                deviation_percent = (deviation / (sum(cycle_1_avg) / len(cycle_1_avg))) * 100 if sum(cycle_1_avg) / len(cycle_1_avg) != 0 else 0
                deviation_percent_rounded = round(deviation_percent, 2)  # Round to 2 decimal places
                deviation_percent_str = f"{deviation_percent_rounded}%"  # Add % sign
                ws_deviation.cell(row=r_idx, column=10 + idx, value=deviation)
                ws_deviation.cell(row=r_idx, column=14 + idx, value=deviation_percent_str)

    return wb


def main():
    keywords = read_keywords_from_file('keywords.txt')
    
    
    folder_path_extended = 'C:\\Users\\vyank\\OneDrive\\Desktop\\Performance-auto\\Files_Extended'

    
    csv_files_extended = [os.path.join(folder_path_extended, file) for file in os.listdir(folder_path_extended) if file.endswith('.csv')]
    combined_data = read_and_process_csv(csv_files_extended, keywords)
    wb = create_excel_workbook_extended(combined_data, [os.path.basename(file) for file in csv_files_extended])
    output_file_extended = "extended-module-result.xlsx"
    wb.save(output_file_extended)
    print(f'Processed data saved to extended-module-result.xlsx')

        
if __name__ == "__main__":
    main()